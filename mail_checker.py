import time
import random
import zipfile
import warnings
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
from config import *


class Navigation:
    def __init__(self, driver, wait):
        self.driver = driver
        self.wait = wait


    def find_mail_input_apple(self, input_mail):
        try:
            self.wait.until(EC.element_to_be_clickable((By.XPATH, input_mail)))
            return True
        except Exception as error:
            print(f"Error in find_mail_input_apple func\nPossible connection error (503).\nTry enabling proxy servers.")
            return False


    def input_mail_in_form_apple(self, input_mail, password_field, password_field_2, email):
        try:
            input_field = self.wait.until(EC.element_to_be_clickable((By.XPATH, input_mail)))
            input_field.clear()
            time.sleep(random.uniform(1, 1.5))
            input_field.send_keys(email)
            time.sleep(random.uniform(1, 1.5))
            click_on_password_field = self.wait.until(EC.element_to_be_clickable((By.XPATH, password_field))).click()
            time.sleep(random.uniform(2, 2.5))
            return True
        except Exception as error:
            print(f"Error in input_mail_in_form_apple func\n")
            return False


    def check_error_msg_apple(self, error_msg):
        try:
            self.wait.until(EC.element_to_be_clickable((By.XPATH, error_msg)))
            return True
        except Exception as error:
            return False


    def press_welcome_buttons_spotify(self, welcome_button_1, welcome_button_2):
        try:
            self.wait.until(EC.element_to_be_clickable((By.XPATH, welcome_button_1))).click()
            time.sleep(random.uniform(0.4, 0.5))
            self.wait.until(EC.element_to_be_clickable((By.XPATH, welcome_button_2))).click()
            return True
        except Exception as error:
            return False


    def find_mail_input_spotify(self, input_mail):
        try:
            self.wait.until(EC.element_to_be_clickable((By.XPATH, input_mail)))
            return True
        except Exception as error:
            print(f"Error in find_mail_input_apple func\n")
            return False


    def input_mail_in_form_spotify(self, input_mail, enter_button, email):
        try:
            input_field = self.wait.until(EC.element_to_be_clickable((By.XPATH, input_mail)))
            self.driver.execute_script("arguments[0].value = '';", input_field)
            time.sleep(random.uniform(0.5, 0.8))
            input_field.send_keys(email)
            time.sleep(random.uniform(0.7, 0.9))
            return True
        except Exception as error:
            print(f"Error in input_mail_in_form_apple func\n")
            return False


    def check_error_msg_spotify(self, error_msg):
        try:
            self.wait.until(EC.element_to_be_clickable((By.XPATH, error_msg)))
            return True
        except Exception as error:
            return False


    def click_om_empty_place_spotify(self, empty_space):
        try:
            self.wait.until(EC.element_to_be_clickable((By.XPATH, empty_space))).click()
            return True
        except Exception as error:
            return False


def create_new_excel():
    workbook = Workbook()
    sheet = workbook.active

    sheet.append(['EMAIL', 'APPLE', 'SPOTIFY'])

    now = datetime.now()
    current_excel_name = now.strftime("%d_%m_%y_%H.%M")
    workbook.save("out/{}.xlsx".format(current_excel_name))
    return "out/{}.xlsx".format(current_excel_name)


def insert_email(workbook_name, email):
    workbook = openpyxl.load_workbook(workbook_name)
    sheet = workbook.active
    sheet.append([email])
    workbook.save(workbook_name)


def find_row_by_email(workbook_name, email):
    workbook = openpyxl.load_workbook(workbook_name)
    sheet = workbook.active
    for row, cell in enumerate(sheet['A'], start=1):
        if cell.value == email:
            return row
    return None


def colorize_column_apple(workbook_name, email, apple_status):
    workbook = openpyxl.load_workbook(workbook_name)
    sheet = workbook.active
    row = find_row_by_email(workbook_name, email)
    if row is not None:
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        apple_cell = sheet.cell(row=row, column=2)
        if apple_status:
            apple_cell.fill = red_fill
            apple_cell.value = "NOT REGISTERED"
        else:
            apple_cell.fill = green_fill
            apple_cell.value = "REGISTERED"
        workbook.save(workbook_name)


def colorize_column_spotify(workbook_name, email, spotify_status):
    workbook = openpyxl.load_workbook(workbook_name)
    sheet = workbook.active
    row = find_row_by_email(workbook_name, email)
    if row is not None:
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        spotify_cell = sheet.cell(row=row, column=3)
        if spotify_status:
            spotify_cell.fill = red_fill
            spotify_cell.value = "NOT REGISTERED"
        else:
            spotify_cell.fill = green_fill
            spotify_cell.value = "REGISTERED"
        workbook.save(workbook_name)


def get_chromedriver(proxy_ip, proxy_port, proxy_login, proxy_pass, use_proxy=False, user_agent=None):
    manifest_json = """
    {
        "version": "1.0.0",
        "manifest_version": 2,
        "name": "Chrome Proxy",
        "permissions": [
            "proxy",
            "tabs",
            "unlimitedStorage",
            "storage",
            "<all_urls>",
            "webRequest",
            "webRequestBlocking"
        ],
        "background": {
            "scripts": ["background.js"]
        },
        "minimum_chrome_version":"22.0.0"
    }
    """

    background_js = """
    var config = {
            mode: "fixed_servers",
            rules: {
            singleProxy: {
                scheme: "http",
                host: "%s",
                port: parseInt(%s)
            },
            bypassList: ["localhost"]
            }
        };

    chrome.proxy.settings.set({value: config, scope: "regular"}, function() {});

    function callbackFn(details) {
        return {
            authCredentials: {
                username: "%s",
                password: "%s"
            }
        };
    }

    chrome.webRequest.onAuthRequired.addListener(
                callbackFn,
                {urls: ["<all_urls>"]},
                ['blocking']
    );
    """ % (proxy_ip, proxy_port, proxy_login, proxy_pass)


    options = webdriver.ChromeOptions()
    options.binary_location = chrome_path
    # options.binary_location = '/usr/bin/google-chrome'
    # options.add_argument('--headless=new')
    # options.add_argument("--disable-gpu")
    # options.add_argument("--no-sandbox")
    # service = webdriver.chrome.service.Service('/root/telegram_bot/chromedriver')
    service = webdriver.chrome.service.Service()
    warnings.filterwarnings("ignore", category=DeprecationWarning)

    if use_proxy:
        pluginfile = 'proxy_auth_plugin.zip'

        with zipfile.ZipFile(pluginfile, 'w') as zp:
            zp.writestr("manifest.json", manifest_json)
            zp.writestr("background.js", background_js)
        options.add_extension(pluginfile)
    if user_agent:
        options.add_argument('--user-agent=%s' % user_agent)
    browser = webdriver.Chrome(service=service, options=options)
    return browser


def spotify_website(proxy_use, mail_list, excel_table, xpath, proxy_file):
    while True:
        if proxy_use == True:
            with open(proxy_file, 'r') as file:
                x = file.read()
            proxy_lst = x.split('\n')
            current_proxy = random.choice(proxy_lst)
            ip, port, login, password = current_proxy.split(':')
            driver = get_chromedriver(ip, port, login, password, use_proxy=proxy_use)
        else:
            driver = webdriver.Chrome()

        wait = WebDriverWait(driver, 2)

        driver.maximize_window()
        driver.get("https://www.spotify.com/md-ru/signup")
        time.sleep(1)

        navigation = Navigation(driver, wait)
        press_welcome_buttons = navigation.press_welcome_buttons_spotify(xpath['welcome_spotify_button_1'], xpath['welcome_spotify_button_2'])
        find_mail_input_field = navigation.find_mail_input_spotify(xpath['input_mail_spotify'])
        if find_mail_input_field == True:
            break
        else:
            time.sleep(5)
            driver.close()

    find_mail_input_field = navigation.find_mail_input_spotify(xpath['input_mail_spotify'])
    if find_mail_input_field == True:
        time.sleep(random.uniform(0.1, 0.25))
        for email in mail_list:
            input_email_address = navigation.input_mail_in_form_spotify(xpath['input_mail_spotify'], xpath['enter_button'], email)
            if input_email_address == True:
                insert_email(excel_table, email)
                time.sleep(random.uniform(0.1, 0.25))
                navigation.click_om_empty_place_spotify(xpath['empty_place_spotify'])
                time.sleep(random.uniform(0.1, 0.25))
                valid_checking = navigation.check_error_msg_spotify(xpath['error_msg_spotify'])
                if valid_checking == True:
                    colorize_column_spotify(excel_table, email, False)
                    print(f"{email} - REGISTERED on spotify")
                else:
                    colorize_column_spotify(excel_table, email, True)
                    print(f"{email} - NOT REGISTERED on spotify")
                time.sleep(random.uniform(0.1, 0.25))
            time.sleep(random.uniform(1, 1.5))
    driver.close()


def apple_website(proxy_use, mail_list, excel_table, xpath, proxy_file):
    count = 0
    for email in mail_list:
        while True:
            if proxy_use == True:
                with open(proxy_file, 'r') as file:
                    x = file.read()

                proxy_lst = x.split('\n')
                proxy_lst = [i for i in proxy_lst if len(i) > 0]

                proxy_ip, proxy_port, proxy_login, proxy_password = proxy_lst[count % len(proxy_lst)].split(':')
                driver = get_chromedriver(proxy_ip, proxy_port, proxy_login, proxy_password, use_proxy=proxy_use)
                count += 1
            else:
                driver = webdriver.Chrome()

            wait = WebDriverWait(driver, 2)

            driver.maximize_window()
            driver.get("https://appleid.apple.com/account")
            time.sleep(2)
            navigation = Navigation(driver, wait)
            find_mail_input_field = navigation.find_mail_input_apple(xpath['input_mail_apple'])
            if find_mail_input_field == True:
                break
            else:
                time.sleep(2)
                driver.close()

        find_mail_input_field = navigation.find_mail_input_apple(xpath['input_mail_apple'])
        if find_mail_input_field == True:
            time.sleep(random.uniform(0.1, 0.25))
            input_email_address = navigation.input_mail_in_form_apple(xpath['input_mail_apple'],xpath['password_field_apple'], xpath['password_field_apple_2'], email)
            time.sleep(random.uniform(0.1, 0.25))
            if input_email_address == True:
                time.sleep(random.uniform(0.1, 0.25))

                insert_email(excel_table, email)
                valid_checking = navigation.check_error_msg_apple(xpath['error_msg_apple'])
                if valid_checking == True:
                    colorize_column_apple(excel_table, email, False)
                    print(f"{email} - REGISTERED on apple")
                elif valid_checking == False:
                    colorize_column_apple(excel_table, email, True)
                    print(f"{email} - NOT REGISTERED on apple")
            time.sleep(random.uniform(1, 1.5))
        driver.close()


def multiple_websites(proxy_use, mail_list, excel_table, xpath, proxy_file):
    websites = ['https://appleid.apple.com/account', 'https://www.spotify.com/md-ru/signup']


    for email in mail_list:
        insert_email(excel_table, email)

    count_websites = 0
    for website in websites:
        if count_websites == 0:
            count = 0
            for email in mail_list:
                while True:
                    if proxy_use == True:
                        with open(proxy_file, 'r') as file:
                            x = file.read()

                        proxy_lst = x.split('\n')
                        proxy_lst = [i for i in proxy_lst if len(i) > 0]

                        proxy_ip, proxy_port, proxy_login, proxy_password = proxy_lst[count % len(proxy_lst)].split(':')
                        driver = get_chromedriver(proxy_ip, proxy_port, proxy_login, proxy_password,use_proxy=proxy_use)
                        count += 1
                    else:
                        driver = webdriver.Chrome()

                    wait = WebDriverWait(driver, 2)

                    driver.maximize_window()
                    driver.get('https://appleid.apple.com/account')
                    time.sleep(2)
                    navigation = Navigation(driver, wait)
                    find_mail_input_field = navigation.find_mail_input_apple(xpath['input_mail_apple'])
                    if find_mail_input_field == True:
                        break
                    else:
                        time.sleep(2)
                        driver.close()

                find_mail_input_field = navigation.find_mail_input_apple(xpath['input_mail_apple'])
                if find_mail_input_field == True:
                    time.sleep(random.uniform(0.1, 0.25))
                    input_email_address = navigation.input_mail_in_form_apple(xpath['input_mail_apple'], xpath['password_field_apple'], xpath['password_field_apple_2'], email)
                    time.sleep(random.uniform(0.1, 0.25))
                    if input_email_address == True:
                        time.sleep(random.uniform(0.1, 0.25))
                        valid_checking = navigation.check_error_msg_apple(xpath['error_msg_apple'])
                        if valid_checking == True:
                            colorize_column_apple(excel_table, email, False)
                            print(f"{email} - REGISTERED on apple")
                        elif valid_checking == False:
                            colorize_column_apple(excel_table, email, True)
                            print(f"{email} - NOT REGISTERED on apple")
                    time.sleep(random.uniform(1, 1.5))
                driver.close()
        elif count_websites == 1:
            while True:
                if proxy_use == True:
                    with open(proxy_file, 'r') as file:
                        x = file.read()

                    proxy_lst = x.split('\n')
                    current_proxy = random.choice(proxy_lst)
                    ip, port, login, password = current_proxy.split(':')
                    driver = get_chromedriver(ip, port, login, password, use_proxy=proxy_use)
                else:
                    driver = webdriver.Chrome()

                wait = WebDriverWait(driver, 2)
                driver.maximize_window()
                navigation = Navigation(driver, wait)
                driver.get('https://www.spotify.com/md-ru/signup')
                press_welcome_buttons = navigation.press_welcome_buttons_spotify(xpath['welcome_spotify_button_1'], xpath['welcome_spotify_button_2'])
                find_mail_input_field = navigation.find_mail_input_spotify(xpath['input_mail_spotify'])
                if find_mail_input_field == True:
                    break
                else:
                    time.sleep(2)
                    driver.close()
            press_welcome_buttons = navigation.press_welcome_buttons_spotify(xpath['welcome_spotify_button_1'], xpath['welcome_spotify_button_2'])
            find_mail_input_field = navigation.find_mail_input_spotify(xpath['input_mail_spotify'])
            if find_mail_input_field == True:
                time.sleep(random.uniform(0.1, 0.25))
                for email in mail_list:
                    input_email_address = navigation.input_mail_in_form_spotify(xpath['input_mail_spotify'], xpath['enter_button'], email)
                    if input_email_address == True:
                        time.sleep(random.uniform(0.1, 0.25))
                        navigation.click_om_empty_place_spotify(xpath['empty_place_spotify'])
                        time.sleep(random.uniform(0.1, 0.25))
                        valid_checking = navigation.check_error_msg_spotify(xpath['error_msg_spotify'])
                        if valid_checking == True:
                            colorize_column_spotify(excel_table, email, False)
                            print(f"{email} - REGISTERED on spotify")
                        else:
                            colorize_column_spotify(excel_table, email, True)
                            print(f"{email} - NOT REGISTERED on spotify")
                        time.sleep(random.uniform(0.1, 0.25))
                    time.sleep(random.uniform(1, 1.5))

            driver.close()
        count_websites += 1


def main(apple, spotify, proxy_use, proxy_file, mails_file):
    xpath = {
        'input_mail_apple': '/html/body/div[2]/aid-web/div[2]/div/div/create-app/aid-create/idms-flow/div/div/div/idms-step/div/div/div/div[2]/div/div/div[3]/div/div[1]/div/account-name/div/div/email-input/div/idms-textbox/idms-error-wrapper/div/div/input',
        'error_msg_apple': '/html/body/div[2]/aid-web/div[2]/div/div/create-app/aid-create/idms-flow/div/div/div/idms-step/div/div/div/div[2]/div/div/div[3]/div/div[1]/div/account-name/div/div/email-input/div/idms-textbox/idms-error-wrapper/div/idms-error/div/div',
        'password_field_apple': '/html/body/div[2]/aid-web/div[2]/div/div/create-app/aid-create/idms-flow/div/div/div/idms-step/div/div/div/div[2]/div/div/div[3]/div/div[2]/div/new-password/div/div/password-input/div/input',
        'password_field_apple_2': '/html/body/div[2]/aid-web/div[2]/div/div/create-app/aid-create/idms-flow/div/div/div/idms-step/div/div/div/div[2]/div/div/div[2]/div/div[3]/div/wc-birthday/div/div/div/div/masked-date/div',
        'input_mail_spotify': '/html/body/div[1]/main/main/section/div/form/div/div/div/div[2]/input',
        'welcome_spotify_button_1': '/html/body/div[1]/main/main/section/div/div[2]/div/div/div/footer/a/span[1]',
        'welcome_spotify_button_2': '/html/body/div[3]/div[2]/div/div[1]/div/div[2]/div/button[1]',
        'enter_button': '/html/body/div[1]/main/main/section/div/form/button/span[1]',
        'error_msg_spotify': '/html/body/div[1]/main/main/section/div/form/div/div/div/div[2]/div',
        'empty_place_spotify': '/html/body/div[1]/main/main/section/div/div/div[1]',
    }

    excel_table = create_new_excel()

    mail_set = set()
    with open(mails_file, 'r') as file:
        for mail in file.read().split("\n"):
            if len(mail) > 0:
                mail_set.add(mail)
    mail_list = list(mail_set)

    if apple == 1 and spotify == 0:
        apple_website(proxy_use, mail_list, excel_table, xpath, proxy_file)
    elif spotify == 1 and apple == 0:
        spotify_website(proxy_use, mail_list, excel_table, xpath, proxy_file)
    elif apple == 1 and spotify == 1:
        multiple_websites(proxy_use, mail_list, excel_table, xpath, proxy_file)
    else:
        print("You have not selected any of the modes.")


if __name__ == "__main__":
    start_time = time.time()
    print("The bot has started working.")

    main(apple, spotify, proxy_use, proxy_file, mails_file)

    end_time = time.time()
    duration = end_time - start_time
    duration_in_minutes = round(duration / 60)
    print(f"Bot working time: {duration_in_minutes} minute(s)")
